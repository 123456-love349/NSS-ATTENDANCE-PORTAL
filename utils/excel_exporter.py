"""
utils/excel_exporter.py
Generates professional Excel exports for attendance lists and OCR review results.
Applies conditional formatting, custom colors, auto-column widths, and header freezing.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

def export_attendance_excel(rows, filepath):
    """Generate a highly styled and formatted Excel report of attendance rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"
    
    # 1. Embed Logo if exists
    try:
        from openpyxl.drawing.image import Image as OPImage
        logo_path = os.path.join("static", "img", "nss_logo.jpg")
        if os.path.exists(logo_path):
            img = OPImage(logo_path)
            img.width = 65
            img.height = 65
            ws.add_image(img, 'A1')
    except Exception as e:
        print(f"[Excel Exporter] Logo insert bypassed: {e}")
        
    # 2. Add Title block
    ws['C1'] = "NSS Attendance Management System"
    ws['C1'].font = Font(name="Segoe UI", size=16, bold=True, color="1F4D36")
    
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['C2'] = f"Generated: {now_str} | Total Records: {len(rows)}"
    ws['C2'].font = Font(name="Segoe UI", size=10, italic=True, color="555555")
    
    # Empty space
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 10
    
    # 3. Table Headers
    headers = [
        "Student Name", "Branch", "Section", "CRN", "URN", "Phone", 
        "NSS Volunteer", "Mode", "Event Name", "Date", "Time", 
        "Latitude", "Longitude", "Location Address", "Location Status", 
        "Google Maps", "Risk Score", "Risk Level", "Duplicate?", "Proxy?"
    ]
    
    header_row = 5
    ws.row_dimensions[header_row].height = 28
    
    # Styling variables
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4D36", end_color="1F4D36", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # 4. Write Data
    row_fill_even = PatternFill(start_color="F9FBF9", end_color="F9FBF9", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name=font_family, size=10)
    
    current_row = header_row + 1
    for r in rows:
        ws.row_dimensions[current_row].height = 20
        row_fill = row_fill_even if current_row % 2 == 0 else row_fill_odd
        
        # Split created_at into Date & Time
        created_at = r.get("created_at", "")
        date_val, time_val = "-", "-"
        if created_at:
            parts = created_at.split()
            date_val = parts[0]
            if len(parts) > 1:
                time_val = parts[1].split(".")[0]
                
        # Risk level text
        risk_score = r.get("risk_score", 0)
        risk_level = "Low"
        if risk_score >= 50:
            risk_level = "High"
        elif risk_score >= 20:
            risk_level = "Medium"
            
        vals = [
            r.get("student_name", ""),
            r.get("branch", "-") or "-",
            r.get("section", "-") or "-",
            r.get("crn", "-") or "-",
            r.get("urn", "-") or "-",
            r.get("phone", ""),
            "Yes" if r.get("is_nss_volunteer") == 1 else "No",
            r.get("attendance_mode", "").upper(),
            r.get("event_name", ""),
            date_val,
            time_val,
            r.get("latitude", ""),
            r.get("longitude", ""),
            r.get("location_address", "") or "-",
            r.get("location_status", "captured"),
            r.get("google_maps_link", ""),
            risk_score,
            risk_level,
            "Yes" if r.get("is_duplicate") == 1 else "No",
            "Yes" if r.get("is_proxy_suspected") == 1 else "No",
        ]
        
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            
            # Alignments
            if col_idx in [2, 3, 4, 5, 7, 8, 10, 11, 12, 13, 15, 17, 18, 19, 20]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Formatting links
            if col_idx == 16 and val:
                cell.hyperlink = val
                cell.font = Font(name=font_family, size=10, color="0000FF", underline="single")
                
        current_row += 1
        
    # 5. Auto Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't use logo rows to compute widths
        for cell in col[header_row - 1:]:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
            
        # Give some padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    # Limit width of address col
    ws.column_dimensions['N'].width = 30
    
    # 6. Freeze Panes
    ws.freeze_panes = 'A6'  # Freeze header and title
    
    # 7. Add Filters
    ws.auto_filter.ref = f"A5:T{current_row - 1}"
    
    # 8. Conditional Formatting (Red fills for flags)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(name=font_family, size=10, color='9C0006')
    orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_font = Font(name=font_family, size=10, color='9C6500')
    
    # Rule for duplicate flag (Col S == "Yes")
    ws.conditional_formatting.add(
        f"S6:S{current_row - 1}",
        CellIsRule(operator='equal', formula=['"Yes"'], fill=red_fill, font=red_font)
    )
    # Rule for proxy flag (Col T == "Yes")
    ws.conditional_formatting.add(
        f"T6:T{current_row - 1}",
        CellIsRule(operator='equal', formula=['"Yes"'], fill=red_fill, font=red_font)
    )
    # Rule for suspicious location (Col O == "suspicious")
    ws.conditional_formatting.add(
        f"O6:O{current_row - 1}",
        CellIsRule(operator='equal', formula=['"suspicious"'], fill=orange_fill, font=orange_font)
    )
    
    wb.save(filepath)


def export_ocr_report_excel(upload_row, extracted_rows, filepath):
    """Generate a spreadsheet for the batch OCR upload job reviews."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Validation Report"
    
    # Headers info
    ws['A1'] = "NSS Attendance - OCR Validation Report"
    ws['A1'].font = Font(name="Segoe UI", size=14, bold=True, color="1F4D36")
    ws['A2'] = f"Original File: {upload_row['original_filename']} | Date: {upload_row['created_at']}"
    ws['A2'].font = Font(name="Segoe UI", size=10, italic=True)
    
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 10
    
    headers = [
        "Index", "Extracted Name", "CRN", "URN", "Phone", 
        "Fuzzy Match Roster", "DB Duplicate?", "Validation Status"
    ]
    
    header_row = 4
    ws.row_dimensions[header_row].height = 25
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4D36", end_color="1F4D36", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    data_font = Font(name="Segoe UI", size=10)
    current_row = 5
    for idx, r in enumerate(extracted_rows, 1):
        ws.row_dimensions[current_row].height = 20
        
        # Calculate Validation status
        status = "OK"
        if r["is_duplicate_in_db"] == 1:
            status = "Duplicate check-in"
        elif not r["matched_student_id"]:
            status = "Missing in Roster"
        elif not r["extracted_phone"] or len(r["extracted_phone"]) != 10:
            status = "Invalid Phone Format"
            
        vals = [
            idx,
            r["extracted_name"] or "-",
            r["extracted_crn"] or "-",
            r["extracted_urn"] or "-",
            r["extracted_phone"] or "-",
            "Matched" if r["matched_student_id"] else "Unregistered",
            "Yes" if r["is_duplicate_in_db"] == 1 else "No",
            status
        ]
        
        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        if status != "OK":
            row_fill = PatternFill(start_color="FFECEC", end_color="FFECEC", fill_type="solid")
            
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            
            if col_idx in [1, 3, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        current_row += 1
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col[header_row - 1:])
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    ws.freeze_panes = 'A5'
    wb.save(filepath)
